from pathlib import Path
import sys
import os
import pytest

QtWidgets = pytest.importorskip("PySide6.QtWidgets")
QApplication = QtWidgets.QApplication

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from MOTEUR.scraping import profile_manager as pm
from MOTEUR.scraping import history
from MOTEUR.scraping.widgets.image_widget import ImageScraperWidget


def test_scrape_logs_history(tmp_path, monkeypatch):
    pm.PROFILES_FILE = tmp_path / "profiles.json"
    history.HISTORY_FILE = tmp_path / "history.json"
    history.LAST_USED_FILE = tmp_path / "last.json"
    pm.save_profiles([{"name": "p1", "selector": ".a"}])

    app = QApplication.instance() or QApplication([])
    widget = ImageScraperWidget()
    widget.refresh_profiles()
    widget.set_selected_profile("p1")
    urls_file = tmp_path / "urls.txt"
    urls_file.write_text("http://example.com\nhttp://ex2.com\n")
    widget.file_edit.setText(str(urls_file))
    widget.folder_edit.setText(str(tmp_path))

    monkeypatch.setattr(
        "MOTEUR.scraping.widgets.image_widget.scrape_images",
        lambda url, sel, folder: 5,
    )

    widget._start()

    entries = history.load_history()
    assert len(entries) == 2
    assert entries[0]["url"] == "http://example.com"
    assert entries[1]["url"] == "http://ex2.com"
    assert all(e["profile"] == "p1" and e["images"] == 5 for e in entries)
    widget.close()


def test_export_excel(tmp_path, monkeypatch):
    pm.PROFILES_FILE = tmp_path / "profiles.json"
    pm.save_profiles([{"name": "p1", "selector": ".a"}])

    app = QApplication.instance() or QApplication([])
    widget = ImageScraperWidget()
    widget.export_data = [
        {"URL": "u1", "Variant": "v1", "Image": "img1"},
        {"URL": "u2", "Variant": "v2", "Image": "img2"},
    ]

    out_file = tmp_path / "out.xlsx"

    monkeypatch.setattr(
        "MOTEUR.scraping.widgets.image_widget.QFileDialog.getSaveFileName",
        lambda *a, **k: (str(out_file), "")
    )
    infos = []
    monkeypatch.setattr(
        "MOTEUR.scraping.widgets.image_widget.QMessageBox.information",
        lambda *a, **k: infos.append(1)
    )

    widget._export_excel()

    assert out_file.exists()
    from openpyxl import load_workbook

    wb = load_workbook(out_file)
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    assert data[1] == ("u1", "v1", "img1")
    assert data[2] == ("u2", "v2", "img2")
    widget.close()
